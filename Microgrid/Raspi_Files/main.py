# =============================================================
# File: main.py
# Description: Raspberry Pi Master script 
#              This script uses Serial and I2C communication
#              to connect and control the Atverter and Atinverter 
#              converters and create a user interface to allow
#              for someone to change settings of the system in
#              real time
# Author: [Zach Kwast]
# Created: [2025-01-15]
# =============================================================

from smbus2 import SMBus
import struct
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import serial
import time
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox
import os
import re

ac_addr = 0x08   # Bus address for DC to AC converter
dc_conv = serial.Serial('/dev/ttyUSB0', 9600, timeout=1)
time.sleep(2)
dc_conv.reset_input_buffer()  # clear garbage

bus = SMBus(1)   # Use bus 1 for Raspberry Pi (GPIO 2/3)

# Save path and data headers
file_path = "/home/Ras/Documents/Atinverter/Microgrid/Raspi_Files/Sensor_Logbook.xlsx"
headers = ["Time", "DC Voltage", "DC Current", "AC Voltage", 
           "AC Current", " ", "High DC Voltage", "High DC Current", 
           "Low DC Voltage", "Low DC Current"]

# Setup of Excel file
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sensor Logbook"
    ws.append(headers)
    wb.save(file_path)
    print("Excel File created")
else:
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "Sensor Logbook"
    print("File is here")

def readData(addr):
    data = bus.read_i2c_block_data(addr, 0, 16) # 16 bytes of data
    # Separate into 4 byte chunks
    Volt1 = float(struct.unpack('f', bytes(data[0:4]))[0])
    Amp1 = float(struct.unpack('f', bytes(data[4:8]))[0])
    Volt2 = float(struct.unpack('f', bytes(data[8:12]))[0])
    Amp2 = float(struct.unpack('f', bytes(data[12:16]))[0])
    info = ["", round(Volt1,3), round(Amp1,3), round(Volt2,3), round(Amp2,3)]
    return info

def sortData():
    # Read the serial data 7 times and organize it
    VL = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    VH = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    duty = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    state = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    setpoint = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    IL = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    IH = dc_conv.readline().decode('utf-8', errors='ignore').strip()
    info = ["", VH, IH, VL, IL, duty, state, setpoint]
    return info

# Inital states and start time
dc_state = 0
ac_state = 5
start_time = datetime.now()

def UpdateAtverter():
    try:
        global dc_state
        # get state from GUI drop down
        dc_state = selected_dc_state.get()

        #check the state and send to Atverter
        if dc_state == "IDLE":
            dc_conv.write("tooIdle\n".encode())
            #print("Idle Mode")
        if dc_state == "BUCK":
            dc_conv.write("tooBuck\n".encode())
            #print("Buck Mode")
            setpoint_label.config(text=f"Setpoint: 5 V")
        if dc_state == "BOOST":
            dc_conv.write("toBoost\n".encode())
            #print("Boost Mode")
            setpoint_label.config(text=f"Setpoint: 15 V")
        if dc_state == "CURRENT":
            dc_conv.write("current\n".encode())
            #print("Current Mode")

        # Check if the entry is empty or not
        entry_value = entry.get().strip()  # Get value and remove any extra spaces
        if entry_value != "":  # Only process if there's a value
            setpoint = int(float(entry_value)*1000) 
            setpoint_label.config(text=f"Setpoint: {setpoint/1000} V or A")
        else:
            setpoint = 0  # Default to 0 if no value is provided

        if setpoint != 0: # send a new setpoint if not 0
            dc_conv.write(f"{setpoint}\n".encode())
    
        entry.delete(0, tk.END)  # Clear the box
        root.focus_set()         # Remove focus from the entry field

        # Update GUI label
        dc_state_label.config(text=f"DC State: {dc_state}")
        return None
    except OSError as e:

        # Read and process data
        ac_volt_and_current = readData(ac_addr)
        dc_conv.reset_input_buffer()
        dc_conv.write("Request\n".encode())
        dc_volt_and_current = sortData()
        
        # Combine and save data
        all_values = ac_volt_and_current + dc_volt_and_current
        all_values[0] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append(all_values)
        wb.save(file_path)
        messagebox.showerror(f"Error: {e}")


def UpdateAtinverter():
    try:
        global ac_state
        # Get frequency value from GUI dropdown
        freq = selected_frequency.get()

        # Set state based on drop down value
        if freq == "No PWM":
            ac_state = 1
        elif freq == "60 Hz":
            ac_state = 4
        elif freq == "50 Hz":
            ac_state = 3
        
        bus.write_byte(ac_addr, ac_state)  # Send the byte to the device

        # Update labels on the GUI
        freq_label.config(text=f"Frequency: {freq}")
        ac_state_label.config(text=f"AC State: {ac_state}")
        return None
    except OSError as e:
        # Read and process data
        ac_volt_and_current = readData(ac_addr)
        dc_conv.reset_input_buffer()  # clear garbage
        dc_conv.write("Request\n".encode())
        dc_volt_and_current = sortData()
        
        # Combine and save data
        all_values = ac_volt_and_current + dc_volt_and_current
        all_values[0] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append(all_values)
        wb.save(file_path)
        messagebox.showerror(f"Error: {e}")

def RequestAtverter():

    # Request info from the Atverter and sort it to display
    dc_conv.reset_input_buffer()
    dc_conv.write("Request\n".encode())
    dc_volt_and_current = sortData()
    VH = float(re.findall(r'[-+]?\d*\.?\d+', dc_volt_and_current[1])[0]) / 1000.0
    IH = float(re.findall(r'[-+]?\d*\.?\d+', dc_volt_and_current[2])[0]) / 1000.0
    VL = float(re.findall(r'[-+]?\d*\.?\d+', dc_volt_and_current[3])[0]) / 1000.0
    IL= float(re.findall(r'[-+]?\d*\.?\d+', dc_volt_and_current[4])[0]) / 1000.0

    dc_info_label.config(text=f"VH: {round(VH, 3)}V, IH: {round(IH, 3)}A, "
                              f"VL: {round(VL, 3)}V, IL: {round(IL, 3)}A")
    time.sleep(0.5)
    return None

def RequestAtinverter():

    # Get info from the Atinverter and show it on the GUI
    try:
        ac_volt_and_current = readData(ac_addr)
        ac_info_label.config(text=f"DC side: {ac_volt_and_current[1]}V, {ac_volt_and_current[2]}A  " 
                              f"AC side: {ac_volt_and_current[3]}V, {ac_volt_and_current[4]}A")
    except OSError as e:
        print(f"RequestAtinverter failed: {e}")
        messagebox.showerror("I2C Error", f"Failed to read AC data: {e}")
    time.sleep(0.5)
    return None

def PowerDownAtverter():

    global dc_state
    dc_state = 0

    dc_conv.write("tooIdle\n".encode())

    dc_state_label.config(text=f"DC State: {dc_state}")
    return None

def PowerDownAtinverter():

    global ac_state
    ac_state = 5

    bus.write_byte(ac_addr, ac_state) 

    ac_state_label.config(text=f"AC State: {ac_state}")
    return None

def PowerCycle():
    global ac_state
    ac_state = 2

    bus.write_byte(ac_addr, ac_state) 

    ac_state_label.config(text=f"AC State: {ac_state}")

    return None

def check_time_elapsed():
    global start_time
    now = datetime.now()
    elapsed = (now - start_time).total_seconds()

    # Save every 5 minutes (300 seconds)
    if elapsed >= 300:
        try:
            # Read and process data
            ac_volt_and_current = readData(ac_addr)
            dc_conv.reset_input_buffer()  # clear garbage
            dc_conv.write("Request\n".encode())
            dc_volt_and_current = sortData()
            
            # Combine and save data
            all_values = ac_volt_and_current + dc_volt_and_current
            all_values[0] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append(all_values)
            wb.save(file_path)
            print("Excel saved")
            
            # Reset the start time for the next 60-second interval
            start_time = datetime.now()

        except Exception as e:
            print(f"Error during timed save: {e}")
    
    # Repeat the check every 1000 milliseconds (1 second)
    root.after(1000, check_time_elapsed)

# Main GUI window
root = tk.Tk()
root.title("Microgrid Status")
root.geometry("400x400")

notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True)

# --- DC Converter Tab ---
dc_frame = tk.Frame(notebook)
notebook.add(dc_frame, text="DC to DC Atverter")

dc_inner = tk.Frame(dc_frame)
dc_inner.pack(expand=True)

tk.Label(dc_inner, text="DC Converter State:").pack(pady=5)
selected_dc_state = tk.StringVar(root, value="IDLE")
tk.OptionMenu(dc_inner, selected_dc_state, "IDLE", "BUCK", "BOOST", "CURRENT").pack(pady=5)

tk.Label(dc_inner, text="Setpoint (Vout):").pack(pady=5)
entry = tk.Entry(dc_inner)
entry.pack(pady=5)

setpoint_label = tk.Label(dc_inner, text="Setpoint:")
setpoint_label.pack(pady=5)

dc_state_label = tk.Label(dc_inner, text="DC State: 0")
dc_state_label.pack(pady=5)

tk.Button(dc_inner, text="Request Info", command=RequestAtverter).pack(fill='x', padx=20, pady=5)
tk.Button(dc_inner, text="Shutoff Atverter", command=PowerDownAtverter).pack(fill='x', padx=20, pady=5)
tk.Button(dc_inner, text="Apply", command=UpdateAtverter).pack(fill='x', padx=20, pady=5)

dc_info_label = tk.Label(dc_inner, text="None")
dc_info_label.pack(pady=5)

# --- AC Converter Tab ---
ac_frame = tk.Frame(notebook)
notebook.add(ac_frame, text="DC to AC Atinverter")

ac_inner = tk.Frame(ac_frame)
ac_inner.pack(expand=True)

tk.Label(ac_inner, text="AC Frequency:").pack(pady=5)
selected_frequency = tk.StringVar(root, value="60 Hz")
tk.OptionMenu(ac_inner, selected_frequency, "No PWM", "50 Hz", "60 Hz").pack(pady=5)

freq_label = tk.Label(ac_inner, text="Frequency: 60 Hz")
freq_label.pack(pady=5)

ac_state_label = tk.Label(ac_inner, text="AC State: 0")
ac_state_label.pack(pady=5)

tk.Button(ac_inner, text="Request Info", command=RequestAtinverter).pack(fill='x', padx=20, pady=5)
tk.Button(ac_inner, text="Power Cycle AC Gate Driver", command=PowerCycle).pack(fill='x', padx=20, pady=5)
tk.Button(ac_inner, text="Shutoff Atinverter", command=PowerDownAtinverter).pack(fill='x', padx=20, pady=5)
tk.Button(ac_inner, text="Apply", command=UpdateAtinverter).pack(fill='x', padx=20, pady=5)

ac_info_label = tk.Label(ac_inner, text="None")
ac_info_label.pack(pady=5)

# Start periodic data save and GUI updates
check_time_elapsed()
root.mainloop()